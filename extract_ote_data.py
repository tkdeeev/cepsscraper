"""
OTE Excel Data Extractor for E-BRIX
====================================
Extracts key datasets from OTE Annual Market Report Excel files
into clean CSVs for dashboard visualization.

Output CSVs (in extracted_data/):
  - dam_prices.csv        -- hourly DAM spot prices
  - im_prices.csv         -- hourly intraday market prices
  - re_positive.csv       -- positive regulation energy (RE+ AC)
  - re_negative.csv       -- negative regulation energy (RE- AC)
  - imbalances.csv        -- system imbalance settlement prices
  - gas_prices.csv        -- daily gas spot prices
  - dam_indexes.csv       -- daily DAM base/peak/offpeak indexes

Usage:
  python extract_ote_data.py
"""

import os
import csv
import re
from datetime import datetime, date
from typing import Optional

import openpyxl

# --- Configuration -----------------------------------------------------------

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ote_data")
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "extracted_data")

# Electricity market files (xlsx only)
ELEC_FILES = [
    "Annual_market_report_2024_V0.xlsx",
    "Annual_market_report_2025_V0.xlsx",
    "Annual_report_2026_V0_markets_RRD.xlsx",
]

# Gas market files
GAS_FILES = [
    "Annual_market_report_gas_2024_V0.xlsx",
    "Annual_market_report_gas_2025_V0.xlsx",
    "Annual_market_report_gas_2026_V0.xlsx",
]


# --- Helpers -----------------------------------------------------------------

def safe_float(val):
    """Convert a cell value to float, handling None and strings."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", ".").replace("\xa0", "")
    if not s or s == "-":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_date(val):
    """Normalize various date formats to YYYY-MM-DD string."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    # Try DD.MM.YYYY format (gas files)
    m = re.match(r"(\d{2})\.(\d{2})\.(\d{4})", s)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    # Try YYYY-MM-DD format
    m = re.match(r"(\d{4}-\d{2}-\d{2})", s)
    if m:
        return m.group(1)
    return None


def parse_hour(val):
    """
    Parse hour from various formats:
    - Integer: 1, 2, ..., 24
    - Float: 1.0, 2.0, ...
    - String: "00-01", "01-02", ..., "23-24"
    Returns 1-24 integer.
    """
    if val is None:
        return None
    if isinstance(val, (int, float)):
        h = int(val)
        if 1 <= h <= 24:
            return h
        return None
    s = str(val).strip()
    # "00-01" format
    m = re.match(r"^(\d{2})-(\d{2})$", s)
    if m:
        return int(m.group(2))
    # Plain integer
    try:
        h = int(s)
        if 1 <= h <= 24:
            return h
    except ValueError:
        pass
    return None


def iter_data_rows(ws, header_row=6):
    """
    Iterate over data rows in a worksheet, skipping header/empty rows.
    Yields (row_number, cells_list) for each data row.
    """
    row_num = 0
    for row in ws.iter_rows(values_only=True):
        row_num += 1
        if row_num <= header_row:
            continue
        first_cell = str(row[0]).strip().lower() if row[0] else ""
        if not first_cell or first_cell in ("celkem", "total", "sum", "none"):
            continue
        yield row_num, list(row)


def write_csv(filepath, headers, rows):
    """Write rows to CSV."""
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)
    print(f"  OK {os.path.basename(filepath)}: {len(rows)} rows")


# --- Extractors --------------------------------------------------------------

def detect_format(ws):
    """
    Detect the data format based on header row content.
    Returns: '2024_format', 'new_format'
    """
    row_num = 0
    for row in ws.iter_rows(values_only=True, max_row=6):
        row_num += 1
        if row_num == 6:
            headers = [str(c).strip() if c else "" for c in row[:5]]
            if "Hour" in headers[1]:
                return "2024_format"
            if "Period" in headers[1]:
                return "new_format"
    return "unknown"


def extract_dam_prices(wb, year_label, rates):
    """Extract DAM prices from a workbook and populate daily exchange rates."""
    if "DAM" not in wb.sheetnames:
        print(f"  WARNING: No DAM sheet in {year_label}")
        return []

    ws = wb["DAM"]
    fmt = detect_format(ws)
    rows = []

    if fmt == "2024_format":
        # Columns: Day(0) Hour(1) Purchase(2) Sale(3) Saldo(4) Export(5) Import(6)
        #          Volume(7) Price_EUR(8) Price_CZK(9) Rate(10) Amount(11)
        for _, cells in iter_data_rows(ws):
            d = parse_date(cells[0])
            h = parse_hour(cells[1])
            if d and h:
                rate = safe_float(cells[10]) # Rate CZK/EUR
                if d not in rates and rate is not None:
                    rates[d] = rate

                rows.append([
                    d, h,
                    safe_float(cells[8]),   # price
                    safe_float(cells[7]),   # volume_mwh
                    safe_float(cells[4]),   # saldo_mwh
                    safe_float(cells[5]),   # export_mwh
                    safe_float(cells[6]),   # import_mwh
                ])

    elif fmt == "new_format":
        # Columns: Day(0) Period(1) TimeInterval(2) Price_EUR(3) Volume(4)
        #          Purchase(5) Sale(6) Saldo(7) Export(8) Import(9)
        quarter_data = {}
        for _, cells in iter_data_rows(ws):
            d = parse_date(cells[0])
            period = safe_float(cells[1])
            interval = str(cells[2]).strip()
            
            if d and period is not None and interval:
                rate = safe_float(cells[11]) # Rate CZK/EUR
                if d not in rates and rate is not None:
                    rates[d] = rate

                p = safe_float(cells[3])
                v = safe_float(cells[4])
                s = safe_float(cells[7])
                e = safe_float(cells[8])
                i = safe_float(cells[9])

                if ":" in interval: # 15-min data
                    hour = int((period - 1) // 4) + 1
                    key = (d, hour)
                    if key not in quarter_data:
                        quarter_data[key] = {"prices": [], "volumes": [], "saldos": [], "exports": [], "imports": []}
                    if p is not None: quarter_data[key]["prices"].append(p)
                    if v is not None: quarter_data[key]["volumes"].append(v)
                    if s is not None: quarter_data[key]["saldos"].append(s)
                    if e is not None: quarter_data[key]["exports"].append(e)
                    if i is not None: quarter_data[key]["imports"].append(i)
                else: # 60-min data
                    hour = int(period)
                    rows.append([d, hour, p, v, s, e, i])

        for (d, h) in sorted(quarter_data.keys()):
            qd = quarter_data[(d, h)]
            avg_price = sum(qd["prices"]) / len(qd["prices"]) if qd["prices"] else None
            sum_vol = sum(qd["volumes"]) if qd["volumes"] else None
            avg_saldo = sum(qd["saldos"]) / len(qd["saldos"]) if qd["saldos"] else None
            sum_exp = sum(qd["exports"]) if qd["exports"] else None
            sum_imp = sum(qd["imports"]) if qd["imports"] else None
            rows.append([d, h, avg_price, sum_vol, avg_saldo, sum_exp, sum_imp])

    # Sort results
    rows.sort(key=lambda x: (x[0], x[1]))
    return rows


def extract_im_prices(wb, year_label):
    """Extract Intraday Market prices."""
    if "IM (EUR)" not in wb.sheetnames:
        print(f"  WARNING: No IM (EUR) sheet in {year_label}")
        return []

    ws = wb["IM (EUR)"]
    fmt = detect_format(ws)
    rows = []

    if fmt == "2024_format":
        # Columns: Day(0) Hour(1) Volume(2) VolStd(3) VolBlock(4) WAvgPrice(5) ...
        for _, cells in iter_data_rows(ws):
            d = parse_date(cells[0])
            h = parse_hour(cells[1])
            if d and h:
                rows.append([
                    d, h,
                    safe_float(cells[5]),   # weighted_avg_price_eur
                    safe_float(cells[2]),   # volume_mwh
                    safe_float(cells[8]),   # min_price
                    safe_float(cells[9]),   # max_price
                ])

    elif fmt == "new_format":
        quarter_data = {}
        for _, cells in iter_data_rows(ws):
            d = parse_date(cells[0])
            period = safe_float(cells[1])
            interval = str(cells[2]).strip()
            
            if d and period is not None and interval:
                p = safe_float(cells[3])
                v = safe_float(cells[4])
                mn = safe_float(cells[11]) if len(cells) > 11 else None
                mx = safe_float(cells[12]) if len(cells) > 12 else None

                if ":" in interval: # 15-min data
                    hour = int((period - 1) // 4) + 1
                    key = (d, hour)
                    if key not in quarter_data:
                        quarter_data[key] = {"prices": [], "volumes": [], "mins": [], "maxs": []}
                    if p is not None: quarter_data[key]["prices"].append(p)
                    if v is not None: quarter_data[key]["volumes"].append(v)
                    if mn is not None: quarter_data[key]["mins"].append(mn)
                    if mx is not None: quarter_data[key]["maxs"].append(mx)
                else:
                    hour = int(period)
                    rows.append([d, hour, p, v, mn, mx])

        for (d, h) in sorted(quarter_data.keys()):
            qd = quarter_data[(d, h)]
            avg_p = sum(qd["prices"]) / len(qd["prices"]) if qd["prices"] else None
            sum_v = sum(qd["volumes"]) if qd["volumes"] else None
            min_p = min(qd["mins"]) if qd["mins"] else None
            max_p = max(qd["maxs"]) if qd["maxs"] else None
            rows.append([d, h, avg_p, sum_v, min_p, max_p])

    # Sort results
    rows.sort(key=lambda x: (x[0], x[1]))
    return rows


def extract_re_ac(wb, sheet_name, year_label):
    """Extract RE+ AC or RE- AC regulation energy data."""
    if sheet_name not in wb.sheetnames:
        print(f"  WARNING: No {sheet_name} sheet in {year_label}")
        return []

    ws = wb[sheet_name]
    fmt = detect_format(ws)
    rows = []

    if fmt == "2024_format":
        # All years: Day(0) Hour(1) Volume(2) Cost(3)
        for _, cells in iter_data_rows(ws):
            d = parse_date(cells[0])
            h = parse_hour(cells[1])
            if d and h:
                vol = safe_float(cells[2])
                cost = safe_float(cells[3])
                rows.append([d, h, vol, cost])

    elif fmt == "new_format":
        # Columns: Day(0) Period(1) TimeInterval(2) Volume(3) Cost(4)
        quarter_data = {}
        for _, cells in iter_data_rows(ws):
            d = parse_date(cells[0])
            period = safe_float(cells[1])
            interval = str(cells[2]).strip() if len(cells) > 2 else ""
            
            if d and period is not None:
                vol = safe_float(cells[3]) if len(cells) > 3 else None
                cost = safe_float(cells[4]) if len(cells) > 4 else None
                
                if ":" in interval: # 15-min data
                    hour = int((period - 1) // 4) + 1
                    key = (d, hour)
                    if key not in quarter_data:
                        quarter_data[key] = {"volumes": [], "costs": []}
                    if vol is not None: quarter_data[key]["volumes"].append(vol)
                    if cost is not None: quarter_data[key]["costs"].append(cost)
                else: # 60-min data
                    hour = int(period)
                    rows.append([d, hour, vol, cost])
                    
        for (d, h) in sorted(quarter_data.keys()):
            qd = quarter_data[(d, h)]
            sum_vol = sum(qd["volumes"]) if qd["volumes"] else None
            sum_cost = sum(qd["costs"]) if qd["costs"] else None
            rows.append([d, h, sum_vol, sum_cost])

    rows.sort(key=lambda x: (x[0], x[1]))
    return rows


def extract_imbalances(wb, year_label):
    """Extract imbalance settlement data."""
    if "Imbalances" not in wb.sheetnames:
        print(f"  WARNING: No Imbalances sheet in {year_label}")
        return []

    ws = wb["Imbalances"]
    fmt = detect_format(ws)
    rows = []

    if fmt == "2024_format":
        for _, cells in iter_data_rows(ws):
            d = parse_date(cells[0])
            h = parse_hour(cells[1])
            if d and h:
                rows.append([
                    d, h,
                    safe_float(cells[2]),   # system_imbalance_mwh
                    safe_float(cells[3]),   # abs_imbalance_mwh
                    safe_float(cells[9]),   # settlement_price_czk
                    safe_float(cells[10]) if len(cells) > 10 else None,  # counter_price
                ])

    elif fmt == "new_format":
        quarter_data = {}
        for _, cells in iter_data_rows(ws):
            d = parse_date(cells[0])
            period = safe_float(cells[1])
            # Imbalances sheet new format does NOT seem to have 'Time interval' string like DAM if it aligns heavily with Period.
            # But wait, looking at my Imbalance 2025 logs, column 2 is SystemImbalance directly! i.e Day, Period, System_Imbalance
            # Wait! 'Time interval' is NOT in Imbalance 2025.
            # Column 1 is Period (1-96). So it's 100% 15-min intervals! 
            # We don't need to check ":" interval. We just group period 1-4 into hour 1.
            if d and period is not None:
                sys_imb = safe_float(cells[2]) if len(cells) > 2 else None
                abs_imb = safe_float(cells[3]) if len(cells) > 3 else None
                set_price = safe_float(cells[9]) if len(cells) > 9 else None
                cnt_price = safe_float(cells[10]) if len(cells) > 10 else None

                hour = int((period - 1) // 4) + 1
                key = (d, hour)
                if key not in quarter_data:
                    quarter_data[key] = {"sys": [], "abs": [], "set_p": [], "cnt_p": []}
                
                if sys_imb is not None: quarter_data[key]["sys"].append(sys_imb)
                if abs_imb is not None: quarter_data[key]["abs"].append(abs_imb)
                if set_price is not None: quarter_data[key]["set_p"].append(set_price)
                if cnt_price is not None: quarter_data[key]["cnt_p"].append(cnt_price)

        for (d, h) in sorted(quarter_data.keys()):
            qd = quarter_data[(d, h)]
            s_sys = sum(qd["sys"]) if qd["sys"] else None
            s_abs = sum(qd["abs"]) if qd["abs"] else None
            avg_set = sum(qd["set_p"]) / len(qd["set_p"]) if qd["set_p"] else None
            avg_cnt = sum(qd["cnt_p"]) / len(qd["cnt_p"]) if qd["cnt_p"] else None
            rows.append([d, h, s_sys, s_abs, avg_set, avg_cnt])

    rows.sort(key=lambda x: (x[0], x[1]))
    return rows


def extract_gas_prices(wb, year_label):
    """Extract gas intraday market prices."""
    if "IM" not in wb.sheetnames:
        print(f"  WARNING: No IM sheet in {year_label}")
        return []

    ws = wb["IM"]
    rows = []

    # Columns: GasDay(0) Volume(1) WAvgPrice(2) Amount(3) IndexOTE(4)
    for _, cells in iter_data_rows(ws):
        d = parse_date(cells[0])
        if d:
            rows.append([
                d,
                safe_float(cells[2]),   # weighted_avg_price_eur
                safe_float(cells[1]),   # volume_mwh
                safe_float(cells[4]),   # index_ote_eur
            ])

    return rows


def extract_dam_indexes(wb, year_label):
    """Extract DAM daily indexes (base/peak/offpeak)."""
    if "DAM Indexes" not in wb.sheetnames:
        print(f"  WARNING: No DAM Indexes sheet in {year_label}")
        return []

    ws = wb["DAM Indexes"]
    rows = []

    # DAM Indexes has header on row 5
    header_row = 5
    row_num = 0
    for row in ws.iter_rows(values_only=True):
        row_num += 1
        if row_num <= header_row:
            continue
        d = parse_date(row[0])
        if d:
            rows.append([
                d,
                safe_float(row[1]),   # base_load_eur
                safe_float(row[2]),   # peak_load_eur
                safe_float(row[3]),   # offpeak_load_eur
            ])

    return rows


# --- Main --------------------------------------------------------------------

def main():
    print("=" * 60)
    print("OTE Excel Data Extractor for E-BRIX")
    print("=" * 60)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # -- Electricity data --
    daily_rates = {}
    all_dam = []
    all_im = []
    all_re_pos = []
    all_re_neg = []
    all_imbalances = []
    all_dam_indexes = []

    for fname in ELEC_FILES:
        fp = os.path.join(DATA_DIR, fname)
        if not os.path.exists(fp):
            print(f"\nWARNING: File not found: {fname}")
            continue

        print(f"\n>> Processing {fname}...")
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)

        dam = extract_dam_prices(wb, fname, daily_rates)
        all_dam.extend(dam)
        print(f"  DAM: {len(dam)} rows")

        im = extract_im_prices(wb, fname)
        all_im.extend(im)
        print(f"  IM: {len(im)} rows")

        re_pos = extract_re_ac(wb, "RE + AC", fname)
        all_re_pos.extend(re_pos)
        print(f"  RE+AC: {len(re_pos)} rows")

        re_neg = extract_re_ac(wb, "RE - AC", fname)
        all_re_neg.extend(re_neg)
        print(f"  RE-AC: {len(re_neg)} rows")

        imb = extract_imbalances(wb, fname)
        all_imbalances.extend(imb)
        print(f"  Imbalances: {len(imb)} rows")

        idx = extract_dam_indexes(wb, fname)
        all_dam_indexes.extend(idx)
        print(f"  DAM Indexes: {len(idx)} rows")

        wb.close()

    # -- Gas data --
    all_gas = []

    for fname in GAS_FILES:
        fp = os.path.join(DATA_DIR, fname)
        if not os.path.exists(fp):
            print(f"\nWARNING: File not found: {fname}")
            continue

        print(f"\n>> Processing {fname}...")
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)

        gas = extract_gas_prices(wb, fname)
        all_gas.extend(gas)
        print(f"  Gas IM: {len(gas)} rows")

        wb.close()

    # Backfill missing exchange rates
    sorted_dates = sorted(list(set(
        [r[0] for r in all_dam] + 
        [r[0] for r in all_im] + 
        [r[0] for r in all_imbalances] +
        [r[0] for r in all_gas]
    )))
    last_rate = 25.0 # fallback default
    for dt in sorted_dates:
        if dt in daily_rates:
            last_rate = daily_rates[dt]
        else:
            daily_rates[dt] = last_rate

    # Function to create converted outputs
    def save_converted_csvs(currency):
        cur_dir = os.path.join(OUTPUT_DIR, currency)
        os.makedirs(cur_dir, exist_ok=True)
        
        def convert(dt, val, source_cur):
            if val is None: return None
            rate = daily_rates.get(dt, 25.0)
            if currency == "eur" and source_cur == "czk":
                return round(val / rate, 2)
            if currency == "czk" and source_cur == "eur":
                return round(val * rate, 2)
            return val
            
        def process_rows(rows, date_col, source_curs):
            out = []
            for row in rows:
                dt = row[date_col]
                new_row = []
                for idx, cval in enumerate(row):
                    if idx < len(source_curs) and source_curs[idx]:
                        new_row.append(convert(dt, cval, source_curs[idx]))
                    else:
                        new_row.append(cval)
                out.append(new_row)
            return out
            
        write_csv(
            os.path.join(cur_dir, "dam_prices.csv"),
            ["date", "hour", "price", "volume_mwh", "saldo_mwh", "export_mwh", "import_mwh"],
            process_rows(all_dam, 0, [None, None, 'eur', None, None, None, None])
        )
        write_csv(
            os.path.join(cur_dir, "im_prices.csv"),
            ["date", "hour", "price", "volume_mwh", "min_price", "max_price"],
            process_rows(all_im, 0, [None, None, 'eur', None, 'eur', 'eur'])
        )
        write_csv(
            os.path.join(cur_dir, "re_positive.csv"),
            ["date", "hour", "volume_mwh", "cost"],
            process_rows(all_re_pos, 0, [None, None, None, 'czk'])
        )
        write_csv(
            os.path.join(cur_dir, "re_negative.csv"),
            ["date", "hour", "volume_mwh", "cost"],
            process_rows(all_re_neg, 0, [None, None, None, 'czk'])
        )
        write_csv(
            os.path.join(cur_dir, "imbalances.csv"),
            ["date", "hour", "system_imbalance_mwh", "abs_imbalance_mwh", "settlement_price", "counter_price"],
            process_rows(all_imbalances, 0, [None, None, None, None, 'czk', 'czk'])
        )
        write_csv(
            os.path.join(cur_dir, "gas_prices.csv"),
            ["date", "price", "volume_mwh", "index_ote"],
            process_rows(all_gas, 0, [None, 'eur', None, 'eur'])
        )
        write_csv(
            os.path.join(cur_dir, "dam_indexes.csv"),
            ["date", "base_load", "peak_load", "offpeak_load"],
            process_rows(all_dam_indexes, 0, [None, 'eur', 'eur', 'eur'])
        )

    print(f"\n{'-' * 60}")
    print("Writing CSVs to eur/ directory...")
    save_converted_csvs("eur")
    
    print(f"\n{'-' * 60}")
    print("Writing CSVs to czk/ directory...")
    save_converted_csvs("czk")
    
    print(f"\n[DONE] All expected CSVs generated in eur/ and czk/!")
    print("=" * 60)


if __name__ == "__main__":
    main()
