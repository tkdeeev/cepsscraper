"""
Inspect OTE Excel files - focused on key sheets for E-BRIX analysis.
"""
import os
import openpyxl

DATA_DIR = os.path.join(os.path.dirname(__file__), "ote_data")

# Key sheets for E-BRIX analysis
KEY_SHEETS = [
    "DAM", "IM (EUR)", "RE + AC", "RE - AC", "RE +", "RE -",
    "Imbalances", "Imbalances +", "Imbalances -",
    "DAM Indexes", "ERD", "Export", "Import",
    "IM",  # Gas IM
    "Description",
]

def inspect_sheet(wb, sheet_name, max_rows=8):
    """Read first N rows from a sheet."""
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    print(f"\n  --- Sheet: '{sheet_name}' ---")
    row_num = 0
    for row in ws.iter_rows(values_only=True, max_row=max_rows):
        row_num += 1
        cells = []
        for val in row[:12]:  # first 12 cols
            s = str(val) if val is not None else ""
            if len(s) > 35:
                s = s[:32] + "..."
            cells.append(s)
        print(f"    R{row_num}: {cells}")
    
    # Count total rows (sample)
    total = 0
    for _ in ws.iter_rows(values_only=True):
        total += 1
        if total > 10000:
            print(f"  Total rows: >10000")
            return
    print(f"  Total rows: {total}")

def main():
    # Focus on 2024 and 2025 xlsx files 
    target_files = [
        "Annual_market_report_2024_V0.xlsx",
        "Annual_market_report_2025_V0.xlsx",
        "Annual_market_report_gas_2024_V0.xlsx",
        "Annual_market_report_gas_2025_V0.xlsx",
        "Annual_report_2026_V0_markets_RRD.xlsx",
    ]
    
    for fname in target_files:
        fp = os.path.join(DATA_DIR, fname)
        if not os.path.exists(fp):
            print(f"SKIP: {fname}")
            continue
            
        print(f"\n{'='*80}")
        print(f"FILE: {fname}")
        print(f"{'='*80}")
        
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        print(f"All sheets: {wb.sheetnames}")
        
        for sheet in KEY_SHEETS:
            if sheet in wb.sheetnames:
                inspect_sheet(wb, sheet)
        
        wb.close()

if __name__ == "__main__":
    main()
